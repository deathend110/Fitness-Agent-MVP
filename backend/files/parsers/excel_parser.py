from __future__ import annotations

from pathlib import Path
from typing import Any

from openpyxl import load_workbook

TRAINING_COLUMNS = {"date", "exercise", "sets", "reps", "rpe", "weight", "kg", "load"}
MAX_ROWS_PER_SHEET = 20
MAX_TEXT_CHARS = 4000


def parse_excel_file(path: Path) -> dict[str, Any]:
    workbook = load_workbook(path, read_only=True, data_only=True)
    sheet_summaries: list[dict[str, Any]] = []
    training_columns: list[str] = []
    text_blocks: list[str] = []

    try:
        for worksheet in workbook.worksheets:
            rows = list(
                worksheet.iter_rows(
                    min_row=1,
                    max_row=MAX_ROWS_PER_SHEET + 1,
                    values_only=True,
                )
            )
            headers = [_cell_to_text(value) for value in rows[0]] if rows else []
            normalized_headers = [header.strip().lower() for header in headers if header.strip()]
            for header in normalized_headers:
                if header in TRAINING_COLUMNS and header not in training_columns:
                    training_columns.append(header)

            preview_rows = [
                [_cell_to_text(value) for value in row]
                for row in rows[1:]
                if any(value is not None and str(value).strip() for value in row)
            ]
            sheet_summaries.append(
                {
                    "name": worksheet.title,
                    "headers": headers,
                    "rowPreviewCount": len(preview_rows),
                }
            )
            text_blocks.append(_sheet_to_text(worksheet.title, headers, preview_rows))
    finally:
        workbook.close()

    text = "\n\n".join(block for block in text_blocks if block).strip()[:MAX_TEXT_CHARS]
    return {
        "kind": "excel",
        "title": path.stem,
        "summary": _build_summary(sheet_summaries, training_columns),
        "preview": {
            "sheets": sheet_summaries,
            "trainingColumns": training_columns,
            "truncated": len("\n\n".join(text_blocks)) > MAX_TEXT_CHARS,
        },
        "text": text,
    }


def _cell_to_text(value: Any) -> str:
    return "" if value is None else str(value)


def _sheet_to_text(name: str, headers: list[str], rows: list[list[str]]) -> str:
    lines = [f"Sheet: {name}"]
    if headers:
        lines.append("Headers: " + ", ".join(headers))
    for row in rows[:MAX_ROWS_PER_SHEET]:
        lines.append(" | ".join(row))
    return "\n".join(lines)


def _build_summary(sheets: list[dict[str, Any]], training_columns: list[str]) -> str:
    sheet_names = ", ".join(sheet["name"] for sheet in sheets) or "无工作表"
    columns = ", ".join(training_columns) if training_columns else "未识别训练列"
    return f"工作表：{sheet_names}；训练列：{columns}。"
